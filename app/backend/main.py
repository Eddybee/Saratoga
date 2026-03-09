import base64
import io
import json
import logging
import math
import os
import re
import threading
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

import fitz
import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageOps
from pydantic import BaseModel

load_dotenv()

_openai_available = False
_openai_client = None
try:
    import openai as _openai_mod

    _api_key = os.getenv("OPENAI_API_KEY", "")
    if _api_key and not _api_key.startswith("sk-your"):
        _openai_client = _openai_mod.OpenAI(api_key=_api_key)
        _openai_available = True
except ImportError:
    pass

_LLM_MODEL = "gpt-5-mini"
_LLM_PRICING_PER_MILLION = {
    "gpt-5-mini": {"input": 0.25, "output": 2.00},
}
LLM_SEND_MAX_SIDE_PX = 3200
LLM_RETRY_SEND_MAX_SIDE_PX = 2400
EXTRACTION_HEADERS = ["Item", "Dimensions", "Notes"]
_LLM_RESPONSE_FORMAT = {
    "type": "json_schema",
    "json_schema": {
        "name": "plan_extraction",
        "strict": True,
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "headers": {
                    "type": "array",
                    "items": {
                        "type": "string",
                        "enum": EXTRACTION_HEADERS,
                    },
                    "minItems": 3,
                    "maxItems": 3,
                },
                "rows": {
                    "type": "array",
                    "items": {
                        "type": "array",
                        "items": {"type": "string"},
                        "minItems": 3,
                        "maxItems": 3,
                    },
                },
            },
            "required": ["headers", "rows"],
        },
    },
}
PAGE_RENDER_DPI = 144
MAX_TILE_SIDE_PX = 1800
MAX_TILE_AREA_PX = 3_000_000
MAX_TILE_COUNT = 6
TILE_OVERLAP_RATIO = 0.08

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="Saratoga Plan Extractor")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:5174",
        "http://localhost:5175",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path("uploads")
RENDER_DIR = Path("renders")
EXPORT_DIR = Path("exports")
UPLOAD_DIR.mkdir(exist_ok=True)
RENDER_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

CLEANUP_MAX_AGE = 24 * 60 * 60

pdf_store: Dict[str, dict] = {}
extraction_store: List[dict] = []


def _cleanup_old_files() -> None:
    """Delete cached render and export files older than 24 hours."""
    now = time.time()
    cleaned = 0
    for directory in (RENDER_DIR, EXPORT_DIR):
        for file_path in directory.iterdir():
            if file_path.is_file() and (now - file_path.stat().st_mtime) > CLEANUP_MAX_AGE:
                try:
                    file_path.unlink()
                    cleaned += 1
                except OSError:
                    pass
    if cleaned:
        logger.info("Cleanup deleted %s temp files older than 24h", cleaned)


def _schedule_cleanup() -> None:
    _cleanup_old_files()
    timer = threading.Timer(3600, _schedule_cleanup)
    timer.daemon = True
    timer.start()


_schedule_cleanup()


class AOIRequest(BaseModel):
    """Accept percentage coordinates with legacy pixel fields for compatibility."""

    pdf_id: str
    page: int
    pct_x: float = 0.0
    pct_y: float = 0.0
    pct_w: float = 0.0
    pct_h: float = 0.0
    x: Optional[float] = None
    y: Optional[float] = None
    width: Optional[float] = None
    height: Optional[float] = None
    image_width: Optional[float] = None
    image_height: Optional[float] = None
    label: Optional[str] = ""


class ExportRequest(BaseModel):
    extractions: list


class ClientDisconnectedError(Exception):
    pass


def normalize_to_pdf_rect(page: fitz.Page, req: AOIRequest) -> fitz.Rect:
    pdf_rect = page.rect

    if req.pct_w > 0 and req.pct_h > 0:
        x0 = pdf_rect.x0 + (req.pct_x / 100.0) * pdf_rect.width
        y0 = pdf_rect.y0 + (req.pct_y / 100.0) * pdf_rect.height
        x1 = pdf_rect.x0 + ((req.pct_x + req.pct_w) / 100.0) * pdf_rect.width
        y1 = pdf_rect.y0 + ((req.pct_y + req.pct_h) / 100.0) * pdf_rect.height
    elif req.x is not None and req.image_width and req.image_height and req.image_width > 0:
        sx = pdf_rect.width / req.image_width
        sy = pdf_rect.height / req.image_height
        x0 = pdf_rect.x0 + req.x * sx
        y0 = pdf_rect.y0 + req.y * sy
        x1 = pdf_rect.x0 + (req.x + (req.width or 0)) * sx
        y1 = pdf_rect.y0 + (req.y + (req.height or 0)) * sy
    else:
        raise ValueError("No valid coordinates provided")

    x0 = max(pdf_rect.x0, min(x0, pdf_rect.x1))
    y0 = max(pdf_rect.y0, min(y0, pdf_rect.y1))
    x1 = max(pdf_rect.x0, min(x1, pdf_rect.x1))
    y1 = max(pdf_rect.y0, min(y1, pdf_rect.y1))

    clip = fitz.Rect(x0, y0, x1, y1)
    logger.info(
        "Coord mapping pct(%.1f%%,%.1f%% %.1f%%x%.1f%%) -> PDF pts(%.1f,%.1f -> %.1f,%.1f)",
        req.pct_x,
        req.pct_y,
        req.pct_w,
        req.pct_h,
        x0,
        y0,
        x1,
        y1,
    )
    return clip


def _page_pixel_size(page: fitz.Page, dpi: int = PAGE_RENDER_DPI) -> tuple[int, int]:
    scale = dpi / 72.0
    return (
        max(1, int(round(page.rect.width * scale))),
        max(1, int(round(page.rect.height * scale))),
    )


def _page_render_path(pdf_id: str, page_num: int) -> Path:
    return RENDER_DIR / f"{pdf_id}_page_{page_num}_{PAGE_RENDER_DPI}dpi.png"


def _render_page_image(pdf_path: str, page_num: int, output_path: Path) -> Path:
    doc = fitz.open(pdf_path)
    try:
        page = doc[page_num]
        zoom = PAGE_RENDER_DPI / 72.0
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
        pix.save(str(output_path))
        logger.info(
            "Rendered page %s to %s at %s DPI (%sx%s)",
            page_num,
            output_path.name,
            PAGE_RENDER_DPI,
            pix.width,
            pix.height,
        )
        return output_path
    finally:
        doc.close()


def render_clip_from_pdf(pdf_path: str, page_num: int, clip: fitz.Rect, dpi: int = 300) -> Image.Image:
    doc = fitz.open(pdf_path)
    try:
        page = doc[page_num]
        zoom = dpi / 72.0
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip, alpha=False)
        image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        logger.info("Rendered AOI clip %sx%s px at %s DPI", pix.width, pix.height, dpi)
        return image
    finally:
        doc.close()


def recommended_render_dpi(clip: fitz.Rect) -> int:
    short_side_pts = min(clip.width, clip.height)
    if short_side_pts <= 72:
        return 600
    if short_side_pts <= 180:
        return 450
    return 300


def _clip_pixel_size(clip: fitz.Rect, dpi: int) -> tuple[float, float]:
    scale = dpi / 72.0
    return clip.width * scale, clip.height * scale


def build_clip_tiles(clip: fitz.Rect, dpi: int) -> list[fitz.Rect]:
    width_px, height_px = _clip_pixel_size(clip, dpi)
    area_px = width_px * height_px

    # Keep typical schedule/note selections on a single fast pass.
    if width_px <= MAX_TILE_SIDE_PX and height_px <= MAX_TILE_SIDE_PX and area_px <= MAX_TILE_AREA_PX:
        return [clip]

    cols = max(1, math.ceil(width_px / MAX_TILE_SIDE_PX))
    rows = max(1, math.ceil(height_px / MAX_TILE_SIDE_PX))
    area_factor = max(1, math.ceil(math.sqrt(area_px / MAX_TILE_AREA_PX))) if area_px > MAX_TILE_AREA_PX else 1
    cols = max(cols, area_factor)
    rows = max(rows, area_factor)

    while cols * rows > MAX_TILE_COUNT:
        if cols >= rows and cols > 1:
            cols -= 1
        elif rows > 1:
            rows -= 1
        else:
            break

    if cols == 1 and rows == 1:
        return [clip]

    tile_width = clip.width / cols
    tile_height = clip.height / rows
    overlap_x = tile_width * TILE_OVERLAP_RATIO if cols > 1 else 0
    overlap_y = tile_height * TILE_OVERLAP_RATIO if rows > 1 else 0

    tiles: list[fitz.Rect] = []
    for row_index in range(rows):
        for col_index in range(cols):
            x0 = clip.x0 + col_index * tile_width - (overlap_x if col_index > 0 else 0)
            y0 = clip.y0 + row_index * tile_height - (overlap_y if row_index > 0 else 0)
            x1 = clip.x0 + (col_index + 1) * tile_width + (overlap_x if col_index < cols - 1 else 0)
            y1 = clip.y0 + (row_index + 1) * tile_height + (overlap_y if row_index < rows - 1 else 0)
            tiles.append(fitz.Rect(x0, y0, x1, y1))

    logger.info(
        "Split AOI into %s tiles (%s cols x %s rows) for dense region extraction",
        len(tiles),
        cols,
        rows,
    )
    return tiles


async def ensure_request_connected(request: Request) -> None:
    if await request.is_disconnected():
        logger.info("Client disconnected during extraction")
        raise ClientDisconnectedError()


@app.get("/health")
def health():
    return {
        "status": "ok",
        "llm": _openai_available,
        "model": _LLM_MODEL,
    }


@app.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    """Upload a PDF and register page metadata without pre-rendering every page."""
    filename = Path(file.filename or "upload.pdf").name
    if not filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")

    pdf_id = str(uuid.uuid4())[:8]
    pdf_path = UPLOAD_DIR / f"{pdf_id}_{filename}"

    content = await file.read()
    with open(pdf_path, "wb") as file_handle:
        file_handle.write(content)

    logger.info("Uploaded PDF %s -> %s", filename, pdf_path)

    try:
        doc = fitz.open(str(pdf_path))
        try:
            page_sizes = [_page_pixel_size(doc[index], PAGE_RENDER_DPI) for index in range(len(doc))]
        finally:
            doc.close()
    except Exception as exc:
        logger.error("PDF open failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"Failed to read PDF: {str(exc)}")

    pdf_store[pdf_id] = {
        "path": str(pdf_path),
        "filename": filename,
        "pages": len(page_sizes),
        "page_sizes": page_sizes,
    }

    return {
        "pdf_id": pdf_id,
        "filename": filename,
        "pages": len(page_sizes),
        "page_sizes": page_sizes,
    }


@app.get("/pdf/{pdf_id}/page/{page}")
def get_page_image(pdf_id: str, page: int):
    """Render a page image on demand and cache it for reuse."""
    if pdf_id not in pdf_store:
        raise HTTPException(status_code=404, detail="PDF not found")

    info = pdf_store[pdf_id]
    if page < 0 or page >= info["pages"]:
        raise HTTPException(status_code=400, detail="Invalid page number")

    render_path = _page_render_path(pdf_id, page)
    if not render_path.exists():
        try:
            _render_page_image(info["path"], page, render_path)
        except Exception as exc:
            logger.error("Page render failed: %s", exc)
            raise HTTPException(status_code=500, detail=f"Failed to render page: {str(exc)}")

    return FileResponse(render_path, media_type="image/png")


_LLM_SYSTEM_PROMPT = """You are a construction estimator extracting data from blueprint schedules and plan tables.

Look at this image of a construction plan section. Extract ALL visible data into a strict JSON format with three columns: Item, Dimensions, and Notes.

Rules:
1. Return ONLY valid JSON matching the required schema - no markdown fences, no explanation.
2. The JSON must have exactly this shape:
   {
     "headers": ["Item", "Dimensions", "Notes"],
     "rows": [
       ["item name or description", "measurement/size/qty", "any extra info"],
       ...
     ]
   }
3. "Item" = the thing being described (e.g. "Window A", "2x6 Stud Wall", "Footing F1", "Room 101").
4. "Dimensions" = any measurements, sizes, quantities, or specs (e.g. "3'-0\" x 6'-8\"", "246 SF", "12 EA", "2x6 @ 16\" O.C.").
5. "Notes" = any additional info like material, finish, type, location, remarks, or specs that don't fit Item/Dimensions.
6. Preserve every number exactly as printed - do NOT round or convert.
7. Preserve units (SF, LF, EA, CY, etc.) in the Dimensions column.
8. If a cell is empty or unreadable, use an empty string "".
9. If the image contains a schedule or table with more than 3 columns, map them intelligently:
   - First/name column -> Item
   - Size/qty/measurement columns -> Dimensions (combine with " x " or " | " if multiple)
   - Everything else -> Notes (combine with " | " if multiple)
10. If the image contains plain text (not a table), still parse it into Item/Dimensions/Notes rows.
    For example a note "PROVIDE 2x10 HEADER AT ALL OPENINGS" -> Item: "Header", Dimensions: "2x10", Notes: "At all openings"
11. If the region is unreadable or contains no meaningful data, return {"headers": ["Item", "Dimensions", "Notes"], "rows": []}.
12. Do not add any keys other than "headers" and "rows".
"""


def _pil_to_base64(img: Image.Image, fmt: str = "PNG") -> str:
    buf = io.BytesIO()
    img.save(buf, format=fmt)
    return base64.b64encode(buf.getvalue()).decode()


def optimize_llm_image(pil_img: Image.Image, max_send_side: int = LLM_SEND_MAX_SIDE_PX) -> Image.Image:
    img = pil_img.convert("RGB")
    img = ImageOps.expand(img, border=16, fill="white")

    min_side = min(img.size)
    max_side = max(img.size)
    target_min_side = 900
    target_max_side = 1800

    if min_side > 0 and min_side < target_min_side:
        scale = target_min_side / min_side
        if max_side * scale > target_max_side:
            scale = target_max_side / max_side
        new_size = (
            max(1, int(round(img.width * scale))),
            max(1, int(round(img.height * scale))),
        )
        img = img.resize(new_size, Image.Resampling.LANCZOS)

    max_side = max(img.size)
    if max_side > max_send_side:
        scale = max_send_side / max_side
        new_size = (
            max(1, int(round(img.width * scale))),
            max(1, int(round(img.height * scale))),
        )
        img = img.resize(new_size, Image.Resampling.LANCZOS)

    img = ImageEnhance.Contrast(img).enhance(1.08)
    img = ImageEnhance.Sharpness(img).enhance(1.35)
    return img


def extract_json_object(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)

    decoder = json.JSONDecoder()
    for idx, char in enumerate(cleaned):
        if char != "{":
            continue
        try:
            parsed, _ = decoder.raw_decode(cleaned[idx:])
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            continue

    raise json.JSONDecodeError("No JSON object found", cleaned, 0)


def score_llm_result(headers: list[str], rows: list[list[str]]) -> float:
    confidence = 70.0

    if headers[:3] == EXTRACTION_HEADERS:
        confidence += 5
    if rows:
        confidence += 8

    filled_cells = 0
    total_cells = 0
    well_formed_rows = 0
    for row in rows:
        if isinstance(row, list) and len(row) == 3:
            well_formed_rows += 1
            total_cells += 3
            filled_cells += sum(1 for cell in row if str(cell).strip())

    if rows:
        confidence += min(10, well_formed_rows * 2)
    if total_cells:
        confidence += min(10, (filled_cells / total_cells) * 10)

    return round(max(45.0, min(confidence, 95.0)), 1)


def normalize_llm_rows(rows: Any) -> list[list[str]]:
    normalized_rows: list[list[str]] = []

    for row in rows or []:
        if isinstance(row, dict):
            item = str(row.get("Item") or row.get("item") or "").strip()
            dimensions = str(row.get("Dimensions") or row.get("dimensions") or "").strip()
            notes = str(row.get("Notes") or row.get("notes") or "").strip()
            normalized_rows.append([item, dimensions, notes])
            continue

        if not isinstance(row, (list, tuple)):
            normalized_rows.append(["", "", str(row).strip()])
            continue

        cells = [str(cell).strip() for cell in row]
        if len(cells) < 3:
            cells.extend([""] * (3 - len(cells)))
        elif len(cells) > 3:
            cells = [cells[0], cells[1], " | ".join(cell for cell in cells[2:] if cell)]

        normalized_rows.append(cells[:3])

    return normalized_rows


def combine_llm_results(results: list[dict]) -> dict:
    headers = EXTRACTION_HEADERS
    combined_rows: list[list[str]] = []
    combined_lines: list[str] = []
    fallback_lines: list[str] = []
    seen_rows: set[tuple[str, str, str]] = set()
    total_input_tokens = 0
    total_output_tokens = 0
    total_cost = 0.0
    confidences: list[float] = []

    for result in results:
        confidences.append(float(result.get("avg_confidence", 0)))
        usage = result.get("usage", {})
        total_input_tokens += int(usage.get("input_tokens", 0))
        total_output_tokens += int(usage.get("output_tokens", 0))
        total_cost += float(usage.get("estimated_cost_usd", 0.0))

        tables = result.get("table_data") or []
        if tables and isinstance(tables[0], list):
            for row in normalize_llm_rows(tables[0][1:]):
                if not any(cell.strip() for cell in row):
                    continue
                row_key = tuple(re.sub(r"\s+", " ", cell).strip().lower() for cell in row)
                if row_key in seen_rows:
                    continue
                seen_rows.add(row_key)
                combined_rows.append(row)
                combined_lines.append(" | ".join(cell for cell in row if cell.strip()))
            continue

        fallback_text = _clean_text(result.get("text", ""))
        if fallback_text and fallback_text not in fallback_lines:
            fallback_lines.append(fallback_text)

    if not combined_rows and fallback_lines:
        combined_rows = [["", "", line] for line in fallback_lines]
        combined_lines = fallback_lines[:]

    full_text = "\n".join(combined_lines)
    avg_confidence = round(sum(confidences) / len(confidences), 1) if confidences else 0.0
    method = results[0]["method"] if len(results) == 1 else "llm_gpt5_mini_tiled"

    return {
        "text": full_text,
        "table_data": [[headers] + combined_rows],
        "structured_lines": combined_lines,
        "method": method,
        "avg_confidence": avg_confidence,
        "usage": {
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "estimated_cost_usd": round(total_cost, 6),
        },
    }


def has_meaningful_llm_content(result: dict) -> bool:
    structured_lines = result.get("structured_lines") or []
    if any(str(line).strip() for line in structured_lines):
        return True

    table_data = result.get("table_data") or []
    if table_data and isinstance(table_data[0], list) and len(table_data[0]) > 1:
        for row in table_data[0][1:]:
            if any(str(cell).strip() for cell in row):
                return True

    return bool(_clean_text(result.get("text", "")))


def get_usage_value(usage: Any, *names: str) -> int:
    for name in names:
        value = getattr(usage, name, None)
        if value is not None:
            return int(value)
    return 0


def estimate_llm_cost(model: str, input_tokens: int, output_tokens: int) -> float:
    pricing = _LLM_PRICING_PER_MILLION.get(model)
    if not pricing:
        return 0.0
    return round(
        (input_tokens / 1_000_000) * pricing["input"]
        + (output_tokens / 1_000_000) * pricing["output"],
        6,
    )


def run_llm_extraction(pil_img: Image.Image, label: str = "") -> dict:
    if not _openai_available or not _openai_client:
        raise RuntimeError("OpenAI API is not configured")

    raw_text = ""
    total_input_tokens = 0
    total_output_tokens = 0
    attempt_plans = [
        {
            "max_send_side": LLM_SEND_MAX_SIDE_PX,
            "retry_note": "",
        },
        {
            "max_send_side": LLM_RETRY_SEND_MAX_SIDE_PX,
            "retry_note": (
                " The previous attempt returned no usable content. Return ONLY valid JSON matching "
                "the required schema. If any text is visible, provide a best-effort extraction "
                "instead of an empty response."
            ),
        },
    ]

    for attempt_index, attempt in enumerate(attempt_plans, start=1):
        prepared_img = optimize_llm_image(pil_img, max_send_side=attempt["max_send_side"])
        b64 = _pil_to_base64(prepared_img, "PNG")

        user_msg = "Extract all data from this construction plan image."
        if label:
            user_msg += f' The user labelled this region: "{label}".'
        user_msg += attempt["retry_note"]

        try:
            response = _openai_client.chat.completions.create(
                model=_LLM_MODEL,
                messages=[
                    {"role": "system", "content": _LLM_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": user_msg},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{b64}",
                                    "detail": "high",
                                },
                            },
                        ],
                    },
                ],
                max_completion_tokens=4096,
                response_format=_LLM_RESPONSE_FORMAT,
                temperature=1,
            )

            usage = getattr(response, "usage", None)
            total_input_tokens += get_usage_value(usage, "prompt_tokens", "input_tokens")
            total_output_tokens += get_usage_value(usage, "completion_tokens", "output_tokens")

            raw_text = (response.choices[0].message.content or "").strip()
            logger.info("LLM attempt %s raw response length: %s chars", attempt_index, len(raw_text))

            parsed = extract_json_object(raw_text)
            headers = EXTRACTION_HEADERS
            rows = normalize_llm_rows(parsed.get("rows", []))
            avg_confidence = score_llm_result(headers, rows)
            estimated_cost = estimate_llm_cost(_LLM_MODEL, total_input_tokens, total_output_tokens)

            table_data = [[headers] + rows]
            text_lines = []
            for row in rows:
                parts = [str(cell) for cell in row if str(cell).strip()]
                text_lines.append(" | ".join(parts))

            full_text = "\n".join(text_lines)

            return {
                "text": full_text,
                "table_data": table_data,
                "structured_lines": text_lines,
                "method": "llm_gpt5_mini",
                "avg_confidence": avg_confidence,
                "raw_response": raw_text,
                "usage": {
                    "input_tokens": total_input_tokens,
                    "output_tokens": total_output_tokens,
                    "estimated_cost_usd": estimated_cost,
                },
            }
        except json.JSONDecodeError as exc:
            logger.warning("LLM JSON parse failed on attempt %s: %s", attempt_index, exc)
            if raw_text or attempt_index == len(attempt_plans):
                estimated_cost = estimate_llm_cost(_LLM_MODEL, total_input_tokens, total_output_tokens)
                return {
                    "text": raw_text,
                    "table_data": [[EXTRACTION_HEADERS, ["", "", raw_text]]],
                    "structured_lines": [raw_text] if raw_text else [],
                    "method": "llm_gpt5_mini_raw",
                    "avg_confidence": 60,
                    "raw_response": raw_text,
                    "usage": {
                        "input_tokens": total_input_tokens,
                        "output_tokens": total_output_tokens,
                        "estimated_cost_usd": estimated_cost,
                    },
                }
            logger.info("Retrying extraction after empty LLM response")
        except Exception as exc:
            logger.error("LLM extraction failed on attempt %s: %s", attempt_index, exc)
            raise RuntimeError(f"LLM extraction failed: {str(exc)}")

    estimated_cost = estimate_llm_cost(_LLM_MODEL, total_input_tokens, total_output_tokens)
    return {
        "text": raw_text,
        "table_data": [[EXTRACTION_HEADERS, ["", "", raw_text]]],
        "structured_lines": [raw_text] if raw_text else [],
        "method": "llm_gpt5_mini_raw",
        "avg_confidence": 60,
        "raw_response": raw_text,
        "usage": {
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "estimated_cost_usd": estimated_cost,
        },
    }


def _clean_text(raw: str) -> str:
    if not raw:
        return ""
    text = raw.strip()
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


@app.post("/extract-aoi")
async def extract_aoi(req: AOIRequest, request: Request):
    """Extract text and structured rows from an Area of Interest using the LLM pipeline."""
    if req.pdf_id not in pdf_store:
        raise HTTPException(status_code=404, detail="PDF not found")

    info = pdf_store[req.pdf_id]
    if req.page < 0 or req.page >= info["pages"]:
        raise HTTPException(status_code=400, detail="Invalid page number")

    try:
        doc = fitz.open(info["path"])
        try:
            page = doc[req.page]
            clip = normalize_to_pdf_rect(page, req)
        finally:
            doc.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Coordinate error: {str(exc)}")

    render_dpi = recommended_render_dpi(clip)

    try:
        await ensure_request_connected(request)
        preview_img = render_clip_from_pdf(info["path"], req.page, clip, dpi=min(render_dpi, 144))
    except ClientDisconnectedError as exc:
        raise HTTPException(status_code=499, detail="Extraction cancelled") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Preview render failed: {str(exc)}")

    preview_img.thumbnail((600, 400))
    cropped_preview = _pil_to_base64(preview_img)

    if not _openai_available:
        raise HTTPException(status_code=503, detail="OpenAI Vision is not configured")

    try:
        await ensure_request_connected(request)
        cropped_img = render_clip_from_pdf(info["path"], req.page, clip, dpi=render_dpi)
        result = run_llm_extraction(cropped_img, req.label or "")
        await ensure_request_connected(request)
        raw_text = result["text"]
        cleaned_text = _clean_text(raw_text)
        table_data = result.get("table_data", [])
        structured_lines = result.get("structured_lines", [])
        extraction_method = result["method"]
        avg_confidence = result["avg_confidence"]
        llm_usage = result.get("usage", {})
    except ClientDisconnectedError as exc:
        raise HTTPException(status_code=499, detail="Extraction cancelled") from exc
    except Exception as exc:
        logger.exception("LLM extraction failed")
        raise HTTPException(status_code=502, detail=f"Vision extraction failed: {str(exc)}")

    ext_id = str(uuid.uuid4())[:8]
    extraction = {
        "id": ext_id,
        "pdf_id": req.pdf_id,
        "page": req.page,
        "label": req.label or "",
        "raw_text": raw_text,
        "cleaned_text": cleaned_text,
        "table_data": table_data,
        "structured_lines": structured_lines,
        "avg_confidence": round(avg_confidence, 1),
        "extraction_method": extraction_method,
        "cropped_preview": cropped_preview,
        "llm_usage": llm_usage,
        "region": {
            "pct_x": req.pct_x,
            "pct_y": req.pct_y,
            "pct_w": req.pct_w,
            "pct_h": req.pct_h,
        },
    }

    try:
        await ensure_request_connected(request)
    except ClientDisconnectedError as exc:
        raise HTTPException(status_code=499, detail="Extraction cancelled") from exc

    extraction_store.append(extraction)
    logger.info(
        "Extraction %s: %s chars, method=%s, conf=%s",
        ext_id,
        len(raw_text),
        extraction_method,
        avg_confidence,
    )
    return extraction


@app.get("/extractions")
def get_extractions():
    return extraction_store


@app.delete("/extractions/{ext_id}")
def delete_extraction(ext_id: str):
    global extraction_store
    extraction_store = [extraction for extraction in extraction_store if extraction["id"] != ext_id]
    return {"status": "deleted"}


@app.delete("/extractions")
def clear_extractions():
    global extraction_store
    extraction_store = []
    return {"status": "cleared"}


@app.post("/export-excel")
async def export_excel(req: ExportRequest):
    """Export extractions to a styled Excel file."""
    if not req.extractions:
        raise HTTPException(status_code=400, detail="No extractions to export")

    rows = []
    for ext in req.extractions:
        if ext.get("table_data"):
            for table in ext["table_data"]:
                for row_index, row in enumerate(table):
                    if row_index == 0:
                        continue
                    item = row[0] if len(row) > 0 else ""
                    dims = row[1] if len(row) > 1 else ""
                    notes = row[2] if len(row) > 2 else ""
                    rows.append(
                        {
                            "Label": ext.get("label", ""),
                            "Page": ext.get("page", 0) + 1,
                            "Item": item,
                            "Dimensions": dims,
                            "Notes": notes,
                            "Confidence": ext.get("avg_confidence", 0),
                            "Method": ext.get("extraction_method", ""),
                        }
                    )
        else:
            rows.append(
                {
                    "Label": ext.get("label", ""),
                    "Page": ext.get("page", 0) + 1,
                    "Item": ext.get("cleaned_text") or ext.get("raw_text", ""),
                    "Dimensions": "",
                    "Notes": "",
                    "Confidence": ext.get("avg_confidence", 0),
                    "Method": ext.get("extraction_method", ""),
                }
            )

    df = pd.DataFrame(rows)
    export_path = EXPORT_DIR / f"plan_extractions_{uuid.uuid4().hex[:6]}.xlsx"

    with pd.ExcelWriter(str(export_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extractions")
        ws = writer.sheets["Extractions"]

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            max_len = max(
                len(str(ws.cell(row=row_num, column=col_num).value or ""))
                for row_num in range(1, len(df) + 2)
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    return FileResponse(
        str(export_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="plan_extractions.xlsx",
    )
